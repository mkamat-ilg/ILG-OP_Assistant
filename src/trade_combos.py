from __future__ import annotations
from dataclasses import dataclass
from typing import List, Dict, Any, Optional
import openpyxl

from src.formula_eval import eval_excel_formula

@dataclass
class ComboLine:
    trade: str
    material: str
    material_desc: str
    qty: float
    uom: str
    material_type: str

def _infer_type(material: Any, desc: str) -> str:
    d = (desc or "").upper()
    if "LABOR" in d:
        return "Labor"
    # crude heuristic: some labor numbers start with 8
    try:
        s = str(int(str(material)))
        if s.startswith("8"):
            return "Labor"
    except Exception:
        pass
    return "Sundry"

def generate_associated_lines(
    workbook_path: str,
    trade_key: str,
    installation_material_desc: str,
    gross_qty: float,
    gross_uom: str,
    sheet_name_map: Optional[Dict[str, str]] = None,
) -> List[ComboLine]:
    sheet_name_map = sheet_name_map or {
        "CARPET": "CARPET",
        "LVP": "LVP.EVP",
        "TILE": "WALL TILE",
    }
    tab = sheet_name_map.get(trade_key.upper())
    if not tab:
        raise ValueError(f"No worksheet mapping for trade={trade_key}")

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    if tab not in wb.sheetnames:
        raise ValueError(f"Worksheet '{tab}' not found")

    ws = wb[tab]
    header = [c.value for c in ws[1]]

    def col_idx(label: str) -> int:
        for i, v in enumerate(header):
            if isinstance(v, str) and v.strip().upper() == label.upper():
                return i + 1
        raise ValueError(f"Column '{label}' not found in '{tab}'")

    sap_col = col_idx("SAP MATERIAL/LABOR NUMBER")
    trade_col = col_idx("TRADE DESCRIPTION")
    actual_qty_col = col_idx("ACTUAL QUANTITY")
    uom_col = col_idx("UOM")

    desc_cols = [i+1 for i, v in enumerate(header) if isinstance(v, str) and v.strip().upper() == "MATERIAL/LABOR DESCRIPTION"]
    desc_col = max(desc_cols) if desc_cols else None
    if not desc_col:
        raise ValueError("No MATERIAL/LABOR DESCRIPTION column found")

    actual_col_letter = openpyxl.utils.get_column_letter(actual_qty_col)
    cell_map: Dict[str, float] = {}
    lines: List[ComboLine] = []

    for r in range(2, ws.max_row + 1):
        material = ws.cell(row=r, column=sap_col).value
        trd = ws.cell(row=r, column=trade_col).value
        desc = ws.cell(row=r, column=desc_col).value
        uom = ws.cell(row=r, column=uom_col).value
        aq_cell = ws.cell(row=r, column=actual_qty_col)
        aq = aq_cell.value

        if material is None and trd is None and desc is None and aq is None:
            if lines:
                break
            continue

        is_install_row = isinstance(material, str) and "INSTALL" in material.upper()

        if is_install_row:
            qty = float(gross_qty)
            mat_type = "Installation Material"
            material_out = "Installation Material"
            desc_out = str(installation_material_desc or "")
            uom_out = str(gross_uom or uom or "")
        else:
            qty = None
            if isinstance(aq, (int, float)):
                qty = float(aq)
            elif isinstance(aq, str) and aq.startswith("="):
                qty = eval_excel_formula(aq, cell_map)
            if qty is None:
                qty = 0.0
            mat_type = _infer_type(material, str(desc or ""))
            material_out = "" if material is None else str(material)
            desc_out = "" if desc is None else str(desc)
            uom_out = "" if uom is None else str(uom)

        cell_map[f"{actual_col_letter}{r}"] = float(qty)

        lines.append(
            ComboLine(
                trade=str(trd or trade_key),
                material=material_out,
                material_desc=desc_out,
                qty=float(qty),
                uom=uom_out,
                material_type=mat_type,
            )
        )

    return lines
