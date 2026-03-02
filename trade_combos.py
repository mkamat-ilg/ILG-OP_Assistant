from dataclasses import dataclass
from typing import List, Dict, Any, Optional
import openpyxl

from src.formula_eval import eval_excel_formula

@dataclass
class ComboLine:
    trade: str
    sap_material: str
    material_desc: str
    qty: float
    uom: str
    material_type: str  # Installation Material | Sundry | Labor

def _infer_type(sap: Any, desc: str) -> str:
    d = (desc or "").upper()
    if "LABOR" in d:
        return "Labor"
    try:
        s = int(str(sap))
        if str(s).startswith("8"):
            return "Labor"
    except Exception:
        pass
    return "Sundry"

def generate_associated_lines(
    workbook_path: str,
    trade: str,
    installation_material_desc: str,
    gross_qty: float,
    gross_uom: str,
    sheet_name_map: Optional[Dict[str, str]] = None,
) -> List[ComboLine]:
    """
    trade: expected values like CARPET / LVP / TILE (your app-level naming)
    sheet_name_map: maps app trade -> workbook tab name, e.g.
      {"CARPET":"CARPET", "LVP":"LVP.EVP", "TILE":"WALL TILE"}
    """
    sheet_name_map = sheet_name_map or {
        "CARPET": "CARPET",
        "LVP": "LVP.EVP",
        "TILE": "WALL TILE",
    }
    tab = sheet_name_map.get(trade.upper())
    if not tab:
        raise ValueError(f"No worksheet mapping for trade={trade}")

    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    if tab not in wb.sheetnames:
        raise ValueError(f"Worksheet '{tab}' not found in {workbook_path}")

    ws = wb[tab]

    # Heuristic: for CARPET/LVP/WALL TILE tabs, header is on row 1
    # We’ll locate columns by header labels.
    header = [c.value for c in ws[1]]
    def col_idx(label: str) -> int:
        for i, v in enumerate(header):
            if isinstance(v, str) and v.strip().upper() == label.upper():
                return i + 1
        raise ValueError(f"Column '{label}' not found in '{tab}'")

    sap_col = col_idx("SAP MATERIAL/LABOR NUMBER")
    trade_col = col_idx("TRADE DESCRIPTION")
    actual_qty_col = col_idx("ACTUAL QUANTITY")
    uom_col = col_idx("UOM")

    # the “description” we want is typically the second MATERIAL/LABOR DESCRIPTION column.
    # On these tabs it’s the one after UOM.
    # We’ll find the last occurrence of that header.
    desc_cols = [i+1 for i, v in enumerate(header) if isinstance(v, str) and v.strip().upper() == "MATERIAL/LABOR DESCRIPTION"]
    desc_col = max(desc_cols) if desc_cols else None
    if not desc_col:
        raise ValueError("No MATERIAL/LABOR DESCRIPTION column found")

    # Build cell map for J-column-style references in formulas by tracking computed ACTUAL QUANTITY values.
    # We'll map the ACTUAL QUANTITY column letter + row number -> computed value.
    # This lets formulas like =J2*0.95 work (if the sheet's actual qty is in column J).
    actual_col_letter = openpyxl.utils.get_column_letter(actual_qty_col)
    cell_map: Dict[str, float] = {}

    lines: List[ComboLine] = []

    # First pass: iterate rows with content in SAP/trade columns
    for r in range(2, ws.max_row + 1):
        sap = ws.cell(row=r, column=sap_col).value
        trd = ws.cell(row=r, column=trade_col).value
        desc = ws.cell(row=r, column=desc_col).value
        uom = ws.cell(row=r, column=uom_col).value
        aq_cell = ws.cell(row=r, column=actual_qty_col)
        aq = aq_cell.value

        # stop if we hit a long blank area
        if sap is None and trd is None and desc is None and aq is None:
            # allow occasional blank rows, but exit after a few
            # (simple: break on first fully blank block after we have started collecting)
            if lines:
                break
            continue

        # Identify the workbook's "Installation Material" row
        is_install_row = isinstance(sap, str) and "INSTALL" in sap.upper()

        if is_install_row:
            qty = float(gross_qty)
            material_type = "Installation Material"
            sap_out = "Installation Material"
            desc_out = str(installation_material_desc)
            uom_out = str(gross_uom or uom or "")
        else:
            qty = None
            if isinstance(aq, (int, float)):
                qty = float(aq)
            elif isinstance(aq, str) and aq.startswith("="):
                # evaluate formula using the computed cell_map
                qty = eval_excel_formula(aq, cell_map)
            if qty is None:
                qty = 0.0  # default; user can edit

            material_type = _infer_type(sap, str(desc or ""))
            sap_out = "" if sap is None else str(sap)
            desc_out = "" if desc is None else str(desc)
            uom_out = "" if uom is None else str(uom)

        # Update cell_map for this row (so next rows can reference it)
        cell_map[f"{actual_col_letter}{r}"] = float(qty)

        lines.append(
            ComboLine(
                trade=str(trd or trade),
                sap_material=sap_out,
                material_desc=desc_out,
                qty=float(qty),
                uom=uom_out,
                material_type=material_type,
            )
        )

    return lines
